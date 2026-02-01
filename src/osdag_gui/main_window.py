"""
Main application window for Osdag GUI.
Handles tab management, docking icons, and main window controls.
"""
import osdag_gui.resources.resources_rc

import sys
import os, yaml
from pathlib import Path
from PySide6.QtWidgets import (
    QWidget, QPushButton, QVBoxLayout, QHBoxLayout, QApplication, QFileDialog,
    QMainWindow, QTabBar, QTabWidget, QLabel
)
from PySide6.QtSvgWidgets import QSvgWidget
from PySide6.QtCore import Qt, QSize, QEvent, QTimer, QPoint, QRect
from PySide6.QtGui import QIcon, QGuiApplication, QPixmap, QPainter, QColor

from osdag_gui.ui.windows.home_window import HomeWindow
from osdag_gui.ui.windows.template_page import CustomWindow
from osdag_gui.ui.components.dialogs.custom_messagebox import CustomMessageBox, MessageBoxType

from osdag_gui.data.database.database_config import PROJECT_PATH, ID, update_project_path, delete_project_record
from osdag_gui.data.database.database_config import get_module_function
from osdag_core.Common import (
    KEY_DISP_FINPLATE, KEY_DISP_CLEATANGLE, KEY_DISP_ENDPLATE, KEY_DISP_SEATED_ANGLE,
    KEY_DISP_BCENDPLATE, KEY_DISP_BEAMCOVERPLATE, KEY_DISP_BEAMCOVERPLATEWELD, KEY_DISP_BB_EP_SPLICE,
    KEY_DISP_COLUMNCOVERPLATE, KEY_DISP_COLUMNCOVERPLATEWELD, KEY_DISP_COLUMNENDPLATE,
    KEY_DISP_LAPJOINTBOLTED, KEY_DISP_LAPJOINTWELDED, KEY_DISP_BUTTJOINTBOLTED, KEY_DISP_BUTTJOINTWELDED,
    KEY_DISP_TENSION_BOLTED, KEY_DISP_TENSION_WELDED, KEY_DISP_COMPRESSION_COLUMN, KEY_DISP_STRUT_WELDED_END_GUSSET,
    KEY_DISP_STRUT_BOLTED_END_GUSSET, KEY_DISP_FLEXURE, KEY_DISP_FLEXURE2, KEY_DISP_PLATE_GIRDER_WELDED, KEY_DISP_FLEXURE4,
    KEY_DISP_BASE_PLATE, KEY_MODULE, PATH_TO_DATABASE, get_documents_folder, get_db_header
)
# Backend Class Imports
from osdag_gui.OS_safety_protocols import get_cleanup_coordinator
import openpyxl
import platform
import ctypes
from ctypes import wintypes

# ============= Resize implementation start ===============
# Detect OS
IS_WINDOWS = platform.system() == "Windows"
IS_LINUX = platform.system() == "Linux"

# ---------------- DPI AWARENESS (Windows only) ----------------
if IS_WINDOWS:
    try:
        # Try to set per-monitor DPI awareness (Windows 10, version 1703+)
        # This is the modern way and handles multi-monitor with different DPI correctly
        ctypes.windll.shcore.SetProcessDpiAwareness(2)  # PROCESS_PER_MONITOR_DPI_AWARE
    except:
        try:
            # Fallback for older Windows 10 versions
            ctypes.windll.user32.SetProcessDPIAware()
        except:
            pass

# ---------------- WIN32 CONSTANTS ----------------
if IS_WINDOWS:
    GWL_STYLE = -16

    WS_THICKFRAME = 0x00040000
    WS_SYSMENU = 0x00080000
    WS_MINIMIZEBOX = 0x00020000
    WS_MAXIMIZEBOX = 0x00010000

    WM_NCHITTEST = 0x0084
    WM_NCCALCSIZE = 0x0083
    WM_GETMINMAXINFO = 0x0024

    HTLEFT = 10
    HTRIGHT = 11
    HTTOP = 12
    HTTOPLEFT = 13
    HTTOPRIGHT = 14
    HTBOTTOM = 15
    HTBOTTOMLEFT = 16
    HTBOTTOMRIGHT = 17
    HTCAPTION = 2

    DWMWA_USE_IMMERSIVE_DARK_MODE = 20
    DWMWA_CAPTION_COLOR = 35

BORDER_WIDTH = 8
TITLEBAR_HEIGHT = 40
SNAP_THRESHOLD = 20  # Pixels from edge to trigger snap (Linux only)

# ---------------- WIN32 STRUCTURES ----------------
if IS_WINDOWS:
    class POINT(ctypes.Structure):
        _fields_ = [("x", ctypes.c_long), ("y", ctypes.c_long)]

    class MINMAXINFO(ctypes.Structure):
        _fields_ = [
            ("ptReserved", POINT),
            ("ptMaxSize", POINT),
            ("ptMaxPosition", POINT),
            ("ptMinTrackSize", POINT),
            ("ptMaxTrackSize", POINT),
        ]

    class MONITORINFO(ctypes.Structure):
        _fields_ = [
            ("cbSize", wintypes.DWORD),
            ("rcMonitor", wintypes.RECT),
            ("rcWork", wintypes.RECT),
            ("dwFlags", wintypes.DWORD),
        ]

# ---------------- DWM SHADOW (Windows only) ----------------
def apply_dwm_shadow(hwnd):
    if IS_WINDOWS:
        margins = ctypes.c_int * 4
        m = margins(-1, -1, -1, -1)  # extend frame to whole window
        ctypes.windll.dwmapi.DwmExtendFrameIntoClientArea(hwnd, ctypes.byref(m))

def apply_window_style(hwnd):
    if IS_WINDOWS:
        style = ctypes.windll.user32.GetWindowLongW(hwnd, GWL_STYLE)
        style |= (
            WS_THICKFRAME |
            WS_SYSMENU |
            WS_MINIMIZEBOX |
            WS_MAXIMIZEBOX
        )
        ctypes.windll.user32.SetWindowLongW(hwnd, GWL_STYLE, style)

# ---------------- SNAP PREVIEW OVERLAY (Linux only) ----------------
class SnapPreviewOverlay(QWidget):
    """Semi-transparent overlay showing snap preview"""
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.ToolTip | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAttribute(Qt.WA_TransparentForMouseEvents)
        self.setStyleSheet("background: transparent;")
        
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # Draw semi-transparent blue overlay
        color = QColor(100, 150, 255, 80)
        painter.fillRect(self.rect(), color)
        
        # Draw border
        border_color = QColor(100, 150, 255, 150)
        painter.setPen(border_color)
        painter.drawRect(self.rect().adjusted(0, 0, -1, -1))

# ============= Resize implementation ends ===============

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.main_widget_instance = None
        self.setWindowIcon(QIcon(":/images/osdag_logo.png"))
        self.setCursor(Qt.CursorShape.ArrowCursor)

        # To track the number of tabs for each Module
        # To avoid logger dublicacy
        self.module_count = {
            KEY_DISP_FINPLATE: -1,
            KEY_DISP_CLEATANGLE: -1,
            KEY_DISP_ENDPLATE: -1,
            KEY_DISP_SEATED_ANGLE: -1,

            KEY_DISP_BCENDPLATE: -1,

            KEY_DISP_BEAMCOVERPLATE: -1,
            KEY_DISP_BEAMCOVERPLATEWELD: -1,
            KEY_DISP_BB_EP_SPLICE: -1,

            KEY_DISP_COLUMNCOVERPLATE: -1,
            KEY_DISP_COLUMNCOVERPLATEWELD: -1,
            KEY_DISP_COLUMNENDPLATE: -1,

            KEY_DISP_LAPJOINTBOLTED: -1,
            KEY_DISP_LAPJOINTWELDED: -1,
            KEY_DISP_BUTTJOINTBOLTED: -1,
            KEY_DISP_BUTTJOINTWELDED: -1,

            KEY_DISP_TENSION_BOLTED: -1,
            KEY_DISP_TENSION_WELDED: -1,

            KEY_DISP_COMPRESSION_COLUMN: -1,
            KEY_DISP_STRUT_WELDED_END_GUSSET: -1,
            KEY_DISP_STRUT_BOLTED_END_GUSSET: -1,

            KEY_DISP_FLEXURE: -1,
            KEY_DISP_FLEXURE2: -1,
            KEY_DISP_PLATE_GIRDER_WELDED: -1,
            KEY_DISP_FLEXURE4: -1,

            KEY_DISP_BASE_PLATE: -1,
        }

        screen = QGuiApplication.primaryScreen()
        screen_size = screen.availableGeometry()

        app = QApplication.instance()
        self.theme = app.theme_manager

        screen_width = screen_size.width()
        screen_height = screen_size.height()

        # Calculate window size
        window_width = int(7 * screen_width / 10)
        window_height = int((7 * screen_height) / 8)

        # Set window size
        self.resize(window_width, window_height)

        # Center the window
        x = int((screen_width - window_width) / 2)
        y = int((screen_height - window_height) / 2)

        self.setGeometry(x, y, window_width, window_height)

        # ============= Save initial geometry for restore ===============
        if IS_LINUX:
            self.pre_snap_geometry = QRect(x, y, window_width, window_height)
        # ===============================================================
         
        # ============= Resize implementation start ===============
        # Make the window frameless for custom buttons
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Window)
        
        if IS_WINDOWS:
            self.setAttribute(Qt.WA_DontShowOnScreen, True)  # Hide initially on Windows
        
        # Track dragging and resizing
        self.dragging = False
        self.drag_position = QPoint()
        self.resizing = False
        self.resize_start_pos = QPoint()
        self.resize_start_geometry = QRect()
        self.resize_edges = {'left': False, 'right': False, 'top': False, 'bottom': False}
        
        # Linux-specific: Snap preview overlay
        if IS_LINUX:
            self.snap_overlay = SnapPreviewOverlay()
            self.snap_geometry = None
            self.pre_maximize_geometry = None
            self.pre_snap_geometry = None  # Store geometry before any snap
            self.is_snapped_maximized = False
            self.is_snapped = False  # Track if snapped to any position
            self.setMouseTracking(True)
        # ============= Resize implementation ends ===============
           
        self.current_tab_index = 0 # To keep track of the next tab index
        self.btn_size = QSize(30, 30)

        # Initialize UI first, as sidebar will overlay it
        self.init_ui() # Call init_ui before sidebar creation to ensure main content exists
        self.handle_add_tab("Home")

        # Using QTimer to delay maximizing until after the window is fully initialized
        # Before maximizing, so that when we click on Restore it comes to normal state.
        # QTimer.singleShot(0, self.showMaximized)

        # Ensure correct deletion on close
        self.setAttribute(Qt.WA_DeleteOnClose, True)

    def init_ui(self):
        # Main Vertical Layout for the entire window's *content*
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_v_layout = QVBoxLayout(central_widget)
        main_v_layout.setContentsMargins(1, 0, 1, 1)
        main_v_layout.setSpacing(0)

        # --- Top HBox Layout (Contains logo, tabs, and window control buttons) ---
        self.title_bar = QWidget()
        top_h_layout = QHBoxLayout(self.title_bar)
        top_h_layout.setContentsMargins(0, 0, 0, 0)
        top_h_layout.setSpacing(0)

        #------- Title bar Icons Start----------------------
        icon_label_widget = QWidget()
        icon_label_h_layout = QHBoxLayout(icon_label_widget)
        icon_label_h_layout.setContentsMargins(5, 0, 5, 0)
        icon_label_h_layout.setSpacing(0)

        # SVG Widget (Dummy SVG for demonstration)
        self.svg_widget = QSvgWidget()
        self.svg_widget.load(":/vectors/Osdag_logo.svg")
        self.svg_widget.setFixedSize(18, 18)
        icon_label_h_layout.addWidget(self.svg_widget)

        # Keep a reference for event filtering (double-click to maximize/restore)
        self.icon_label_widget = icon_label_widget
        #------- Title bar Icons End----------------------

        # ============= Resize implementation start ===============
        # Linux: Enable mouse tracking
        if IS_LINUX:
            self.title_bar.setMouseTracking(True)
        # ============= Resize implementation ends ===============

        #------- Tabs layout Start ----------------------
        tabs_h_layout = QHBoxLayout()
        tabs_h_layout.setSpacing(0)
        tabs_h_layout.setContentsMargins(0, 2, 0, 0)

        # QTabBar
        self.tab_bar = QTabBar()
        self.tab_bar.setObjectName("main_tabs")
        self.tab_bar.setExpanding(False)
        self.tab_bar.setTabsClosable(True)
        self.tab_bar.setMovable(False)
        self.tab_bar.tabCloseRequested.connect(self.handle_close_tab)
        # ============= Resize implementation start ===============
        # Linux: Enable mouse tracking
        if IS_LINUX:
            self.tab_bar.setMouseTracking(True)
        # ============= Resize implementation ends ===============
        tabs_h_layout.addWidget(self.tab_bar)
        #------- Tabs layout End ----------------------

        # Helper function to create a styled button
        def create_button(icon_svg, is_close=False):
            btn = QPushButton()
            btn.setFixedSize(self.btn_size)
            btn.setIcon(QIcon(QPixmap.fromImage(QPixmap(icon_svg).toImage())))
            btn.setIconSize(QSize(14, 14))
            if is_close:
                btn.setObjectName("close_button")
            else:
                btn.setObjectName("window_control_button")
            return btn
        
        #------- Control buttons Start ----------------------
        self.minimize_button = create_button(":/vectors/window_minimize_light.svg")
        self.minimize_button.clicked.connect(self.showMinimized)
        
        self.maximize_button = create_button(":/vectors/window_maximize_light.svg")
        self.maximize_button.clicked.connect(self.toggle_maximize_restore)
        
        self.close_button = create_button(":/vectors/window_close_light.svg", is_close=True)
        self.close_button.clicked.connect(self.close_osdag)
        #------- Control buttons Start ----------------------

        # Arrange widgets in title bar layout
        window_control_btn_left = False
        if window_control_btn_left:
            top_h_layout.addWidget(self.close_button)
            top_h_layout.addWidget(self.minimize_button)
            top_h_layout.addWidget(self.maximize_button)
            
            top_h_layout.addLayout(tabs_h_layout)
            
            # Stretch to push Icon to the right
            top_h_layout.addStretch(1)
            top_h_layout.addWidget(icon_label_widget)


        else:
            top_h_layout.addWidget(icon_label_widget)
            top_h_layout.addLayout(tabs_h_layout)

            # Stretch to push buttons to the right
            top_h_layout.addStretch(1)

            top_h_layout.addWidget(self.minimize_button)
            top_h_layout.addWidget(self.maximize_button)
            top_h_layout.addWidget(self.close_button)
        
        # Install event filters for double-click maximize/restore on title widgets
        self.tab_bar.installEventFilter(self)
        self.icon_label_widget.installEventFilter(self)

        self.start_pos = None
        self.start_geometry = None

        # Add top HBox to main VBox
        main_v_layout.addWidget(self.title_bar)

        # QTabWidget
        self.tab_widget = QTabWidget()
        self.tab_widget.tabBar().hide()
        self.tab_widget.setTabsClosable(True) # Allow closing tabs
        self.tab_widget.setMovable(False) # Allow reordering tabs
        self.tab_widget_content = []
        self.tab_widget.tabCloseRequested.connect(self.handle_close_tab)
        
        # ============= Resize implementation start ===============
        if IS_LINUX:
            self.tab_widget.setMouseTracking(True)
        # ============= Resize implementation ends ===============
        
        main_v_layout.addWidget(self.tab_widget)

        # Connect the QTabBar to custom handler
        self.tab_bar.currentChanged.connect(self.handle_tab_change)

        # Ensure initial synchronization
        if self.tab_bar.count() > 0:
            self.tab_widget.setCurrentIndex(self.tab_bar.currentIndex())

    # ============= Resize implementation start ===============
    # WINDOWS: WIN32 NATIVE EVENT PROCESSING - FINAL FIX FOR AERO SNAP
    def nativeEvent(self, eventType, message):
        if not IS_WINDOWS:
            return False, 0
        
        msg = wintypes.MSG.from_address(message.__int__())

        # Tell Windows to NOT draw non-client area (title bar)
        if msg.message == WM_NCCALCSIZE:
            return True, 0

        # Handle WM_GETMINMAXINFO to respect working area (exclude taskbar)
        if msg.message == WM_GETMINMAXINFO:
            info = ctypes.cast(msg.lParam, ctypes.POINTER(MINMAXINFO)).contents
            
            # Get the monitor that the window is on
            hwnd = int(self.winId())
            monitor = ctypes.windll.user32.MonitorFromWindow(
                hwnd, 
                2  # MONITOR_DEFAULTTONEAREST
            )
            
            if monitor:
                monitor_info = MONITORINFO()
                monitor_info.cbSize = ctypes.sizeof(MONITORINFO)
                
                if ctypes.windll.user32.GetMonitorInfoW(monitor, ctypes.byref(monitor_info)):
                    # Use work area (excludes taskbar) instead of full monitor
                    work_area = monitor_info.rcWork
                    
                    # Set maximum size to work area
                    info.ptMaxSize.x = work_area.right - work_area.left
                    info.ptMaxSize.y = work_area.bottom - work_area.top
                    
                    # Set maximum position to work area top-left
                    info.ptMaxPosition.x = work_area.left
                    info.ptMaxPosition.y = work_area.top
            
            return True, 0

        # Handle resizing borders and dragging
        if msg.message == WM_NCHITTEST:
            # This is the key to making it work at all DPI scales
            hwnd = int(self.winId())
            
            # Get cursor position in screen coordinates
            cursor_pos = POINT()
            ctypes.windll.user32.GetCursorPos(ctypes.byref(cursor_pos))
            
            # This handles DPI scaling automatically
            ctypes.windll.user32.ScreenToClient(hwnd, ctypes.byref(cursor_pos))
            
            # Get client rectangle
            client_rect = wintypes.RECT()
            ctypes.windll.user32.GetClientRect(hwnd, ctypes.byref(client_rect))
            
            # Extract position and dimensions
            x_pos = cursor_pos.x
            y_pos = cursor_pos.y
            w = client_rect.right - client_rect.left
            h = client_rect.bottom - client_rect.top
            
            # Border width - use fixed value, ScreenToClient handles DPI
            border = 8
            
            # When maximized or fullscreen, disable resize borders
            if self.isMaximized() or self.isFullScreen():
                border = 0
            
            # Check resize zones
            lx = x_pos < border
            rx = x_pos > w - border
            ty = y_pos < border
            by = y_pos > h - border

            # Return resize handles (corners have priority)
            if lx and ty:
                return True, HTTOPLEFT
            if rx and by:
                return True, HTBOTTOMRIGHT
            if rx and ty:
                return True, HTTOPRIGHT
            if lx and by:
                return True, HTBOTTOMLEFT
            if ty:
                return True, HTTOP
            if by:
                return True, HTBOTTOM
            if lx:
                return True, HTLEFT
            if rx:
                return True, HTRIGHT

            # Title bar dragging check
            # Use Qt's coordinate system for widget hit testing
            from PySide6.QtGui import QCursor
            global_pos = QCursor.pos()
            pos = self.mapFromGlobal(global_pos)
            
            # Check if in title bar area
            if pos.y() <= self.title_bar.height() and pos.y() >= 0:                
                widget_at_pos = QApplication.widgetAt(QCursor.pos())

                # If cursor is over ANY interactive widget, don't treat as caption
                if widget_at_pos is not None:

                    # Allow dragging only on empty title bar areas
                    allowed_drag_widgets = {
                        self.title_bar,
                        self.icon_label_widget,
                        self.svg_widget
                    }

                    if widget_at_pos in allowed_drag_widgets:
                        return True, HTCAPTION

                    # Everything else (tabs, tab close buttons, labels, etc.)
                    return False, 0
                
                # For tab bar, check if over actual tab
                if widget_at_pos == self.tab_bar:
                    tab_bar_pos = self.tab_bar.mapFromGlobal(global_pos)
                    tab_index = self.tab_bar.tabAt(tab_bar_pos)
                    if tab_index == -1:
                        return True, HTCAPTION
                    else:
                        return False, 0
                
                # Allow dragging from any other part of the title bar
                return True, HTCAPTION

        return False, 0

    # LINUX: Helper methods
    def get_resize_region(self, pos):
        """Determine which resize region the mouse is in"""
        w, h = self.width(), self.height()
        
        left = pos.x() <= BORDER_WIDTH
        right = pos.x() >= w - BORDER_WIDTH
        top = pos.y() <= BORDER_WIDTH
        bottom = pos.y() >= h - BORDER_WIDTH
        
        return left, right, top, bottom

    def update_cursor(self, pos):
        """Update cursor based on position (Linux only)"""
        left, right, top, bottom = self.get_resize_region(pos)
        
        if (top and left) or (bottom and right):
            self.setCursor(Qt.SizeFDiagCursor)
        elif (top and right) or (bottom and left):
            self.setCursor(Qt.SizeBDiagCursor)
        elif left or right:
            self.setCursor(Qt.SizeHorCursor)
        elif top or bottom:
            self.setCursor(Qt.SizeVerCursor)
        else:
            self.setCursor(Qt.ArrowCursor)

    def get_snap_geometry(self, global_pos):
        """Calculate snap geometry based on cursor position near screen edges (Linux only)"""
        # Get available screen geometry (excludes taskbar)
        screen = QGuiApplication.primaryScreen().availableGeometry()
        
        x = global_pos.x()
        y = global_pos.y()
        
        # Check if near edges
        near_left = x <= screen.left() + SNAP_THRESHOLD
        near_right = x >= screen.right() - SNAP_THRESHOLD
        near_top = y <= screen.top() + SNAP_THRESHOLD
        near_bottom = y >= screen.bottom() - SNAP_THRESHOLD
        
        # Calculate snap regions
        half_width = screen.width() // 2
        half_height = screen.height() // 2
        
        # Corner snaps (quarter screen)
        if near_top and near_left:
            return QRect(screen.left(), screen.top(), half_width, half_height)
        if near_top and near_right:
            return QRect(screen.left() + half_width, screen.top(), half_width, half_height)
        if near_bottom and near_left:
            return QRect(screen.left(), screen.top() + half_height, half_width, half_height)
        if near_bottom and near_right:
            return QRect(screen.left() + half_width, screen.top() + half_height, half_width, half_height)
        
        # Edge snaps (half screen)
        if near_left:
            return QRect(screen.left(), screen.top(), half_width, screen.height())
        if near_right:
            return QRect(screen.left() + half_width, screen.top(), half_width, screen.height())
        if near_top:
            return screen  # Maximize to available area (respects taskbar)
        
        return None

    def hide_snap_preview(self):
        """Hide snap preview overlay (Linux only)"""
        if IS_LINUX:
            self.snap_overlay.hide()
            self.snap_geometry = None

    def show_snap_preview(self, geometry):
        """Show snap preview overlay (Linux only)"""
        if IS_LINUX:
            if geometry:
                self.snap_overlay.setGeometry(geometry)
                self.snap_overlay.show()
                self.snap_geometry = geometry
            else:
                self.hide_snap_preview()
    # ============= Resize implementation ends ===============

    def paintEvent(self, event):
        if self.theme.is_light():
            self.minimize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_minimize_light.svg").toImage())))
            self.close_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_close_light.svg").toImage())))
            if self.isMaximized():
                self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_restore_light.svg").toImage())))
            else:
                self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_maximize_light.svg").toImage())))

        else:
            self.minimize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_minimize_dark.svg").toImage())))
            self.close_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_close_dark.svg").toImage())))
            if self.isMaximized():
                self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_restore_dark.svg").toImage())))
            else:
                self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_maximize_dark.svg").toImage())))

        super().paintEvent(event)

    def set_maximize_icon(self):
        if self.theme.is_light():
            self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_maximize_light.svg").toImage())))
        else:
            self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_maximize_dark.svg").toImage())))

    def set_restore_icon(self):
        if self.theme.is_light():
            self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_restore_light.svg").toImage())))
        else:
            self.maximize_button.setIcon(QIcon(QPixmap.fromImage(QPixmap(":/vectors/window_restore_dark.svg").toImage())))

    def toggle_maximize_restore(self):
        """Toggles between maximized and normal window states and updates the icon."""
        # ============= Resize implementation start ===============
        if IS_LINUX:
            # If currently in any snapped state (maximize or half/quarter)
            if self.is_snapped or self.isMaximized():
                # Restore to geometry before snap
                restore_geom = self.pre_snap_geometry if self.pre_snap_geometry else self.geometry()
                self.showNormal()
                self.setGeometry(restore_geom)
                
                # Clear snap state
                self.is_snapped = False
                self.is_snapped_maximized = False
                self.set_maximize_icon()
            else:
                # Not snapped - save current geometry and maximize
                self.pre_snap_geometry = self.geometry()
                self.showMaximized()
                self.is_snapped_maximized = False
                self.is_snapped = False
                self.set_restore_icon()
        else:
            # Windows: Simple toggle
            if self.isMaximized():
                self.showNormal()
                self.set_maximize_icon()
            else:
                self.showMaximized()
                self.set_restore_icon()
        # ============= Resize implementation ends ===============

    def add_new_tab(self, module):
        """Helper to add a new tab to QTabWidget."""
        body_widget = QWidget()

        # Create and set layout for body_widget first
        self.main_widget_layout = QHBoxLayout(body_widget)
        self.main_widget_layout.setContentsMargins(0, 0, 0, 0)
        self.main_widget_layout.setSpacing(0)

        # ============= Resize implementation start ===============
        if IS_LINUX:
            body_widget.setMouseTracking(True)
        # ============= Resize implementation ends ===============

        # it initially sets the home on the Tab
        self.open_home_page(module)
        # Widget of the Module
        self.tab_widget_content.append(body_widget)
        self.tab_widget.addTab(body_widget, f"Tab {self.current_tab_index + 1}")
        # Update main_widget_layout to the layout of the new tab's body_widget
        if hasattr(body_widget, 'layout'):
            self.main_widget_layout = body_widget.layout()

    def handle_add_tab(self, module):
        """Handles the 'Add New Tab' button click."""
        self.current_tab_index += 1
        self.tab_bar.addTab("Home") # Add to tab bar
        # Set the newly added tab as current
        self.add_new_tab(module) # Add to tab widget
        
        new_index = self.tab_bar.count() - 1
        self.tab_bar.setCurrentIndex(new_index)
        self.tab_widget.setCurrentIndex(new_index)

        # print("@[New Tab Added]Total Widgets. ", len(QApplication.allWidgets()))

    def handle_tab_change(self, index):
        # Switch the QTabWidget to the new tab
        if index < len(self.tab_widget_content) and index >= 0:
            self.tab_widget.setCurrentIndex(index)

            # Update main_widget_instance to the main widget in the current tab
            body_widget = self.tab_widget_content[index]
            if hasattr(body_widget, 'layout') and body_widget.layout().count() > 0:
                widget_item = body_widget.layout().itemAt(0)
                if widget_item is not None:
                    widget = widget_item.widget()
                    if widget is not None:
                        self.main_widget_instance = widget
            # Update main_widget_layout to the layout of the current tab's body_widget
            if hasattr(body_widget, 'layout'):
                self.main_widget_layout = body_widget.layout()

    # This is triggered by Quit button in Menu bar on template_page
    def close_current_tab(self):
        current_index = self.tab_bar.currentIndex()
        self.handle_close_tab(current_index)

    # General closing function
    def handle_close_tab(self, index)-> bool:
        print(f"[TAB CLOSE] Requested to close tab index: {index}")

        tab_title = self.tab_bar.tabText(index) if index >= 0 else "Module"
        is_last_tab = self.tab_widget.count() == 1
        to_save = self._check_design_done(index)
        module = self._get_template_instance(index)
        
        if to_save and is_last_tab:
            # Check if we're already in close_osdag flow - if so, just close the tab
            if getattr(self, '_closing_tabs', False):
                self._close_tab(index)
            else:
                result = CustomMessageBox(
                    title="Confirm Exit",
                    text=(
                        f"'{tab_title}' is the last tab.\n"
                         "Closing it will exit Osdag.\n"
                        f"Do you want to save your '{tab_title}' design before closing?"
                    ),
                    buttons=["Go to Home", "Save and Exit", "Exit Without Saving", "Cancel"]
                ).exec()
                
                if result == "Save and Exit":
                    # Call Save Function
                    module.saveDesign()
                    # Close tab first, then exit Osdag
                    self._close_tab(index)
                    self.close()
                elif result == "Exit Without Saving":
                    # Close tab first, then exit Osdag
                    self._close_tab(index)
                    self.close()
                elif result == "Go to Home":
                    # Open New Tab & Close This Tab
                    self.handle_add_tab("Home")
                    self._close_tab(index)
                elif result == "Cancel":
                    return False
        
        elif to_save:
            result = CustomMessageBox(
                title="Save Design",
                text=f" Do you want to Save Your '{tab_title}' design before closing?",
                buttons=["Yes", "No"],
                dialogType=MessageBoxType.Warning,
            ).exec()

            if result == "Yes":
                # Call Save Function
                module.saveDesign()
                self._close_tab(index)
            elif result == "No":
                # Close Tab
                self._close_tab(index)
            elif result == "Cancel":
                return False

        elif is_last_tab:
            # Check if we're already in close_osdag flow - if so, just close the tab
            if getattr(self, '_closing_tabs', False):
                self._close_tab(index)
            else:
                options = ["Yes", "No"]
                if tab_title != "Home":
                    options.insert(0, "Go to Home") # Insert at beginning
                # User clicked X on the last tab - ask confirmation
                result = CustomMessageBox(
                    title="Confirm Exit",
                    text=f"'{tab_title}' is the last tab.\nClosing it will exit Osdag.\nDo you really want to close this tab?",
                    buttons=options,
                    dialogType=MessageBoxType.Warning,
                ).exec()

                # Handle result
                if result == "Yes":
                    # CRITICAL: Close the tab FIRST to cleanup CAD, then exit app
                    self._close_tab(index)
                    self.close()  # Close the main window (exit Osdag)
                elif result == "Go to Home":
                    # Open New Tab & Close This Tab
                    self.handle_add_tab("Home")
                    self._close_tab(index)
                elif result == "No":
                    return False
                elif result == "Cancel":
                    return False
        else:
            self._close_tab(index)

        # closed successfully
        return True

    # Check if design is created in the module or not
    def _check_design_done(self, index) -> bool:
        module = self._get_template_instance(index)
        if hasattr(module, 'backend'):
            return module.backend.design_status
        else:
            return False
    
    # It is triggered by quit and close button of main window
    def close_osdag(self):
        # Close all tabs one by one
        while self.tab_bar.count() > 0:
            current_index = self.tab_bar.currentIndex()
            close = self.handle_close_tab(current_index)
            if close is False:
                # If someone cancel to save while closing tabs, then stop closing further tabs
                return
            # Cleanup Coordinator takes some time
            # All tabs closed
            if current_index == 0:
                return
        # Finally the main window closed
        
    def _get_template_instance(self, index) -> object:
        return self.tab_widget_content[index].layout().itemAt(0).widget()

    def clear_layout(self, layout):
        """Properly clear layout with signal disconnection and widget cleanup."""
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                try:
                    widget.setUpdatesEnabled(False)
                    widget.blockSignals(True)
                    widget.hide()
                    
                    # Disconnect specific signals if they exist
                    signals = ['openNewTab', 'downloadDatabase', 'triggerLoadOsi',
                            'openProject', 'openModule', 'cardOpenClicked']
                    
                    for sig in signals:
                        if hasattr(widget, sig):
                            try:
                                getattr(widget, sig).disconnect()
                            except:
                                pass
                    
                    
                    # Use CleanupCoordinator for safe cleanup
                    # This replaces the legacy graveyard/setParent(None) logic
                    from osdag_gui.OS_safety_protocols import get_cleanup_coordinator
                    coordinator = get_cleanup_coordinator()
                    
                    # If the widget has CAD capability, treat it carefully
                    if hasattr(widget, 'cad_widget') and widget.cad_widget:
                        coordinator.cleanup_for_tab_close(widget)
                    else:
                        widget.deleteLater()
                        
                except (RuntimeError, TypeError, Exception) as e:
                    print(f"[WARNING] Error clearing layout widget: {e}")
                    pass
            else:
                sub_layout = item.layout()
                if sub_layout:
                    self.clear_layout(sub_layout)
                    sub_layout.deleteLater()

    def _cleanup_scroll_area(self, scroll_area):
        """Special cleanup for QScrollArea widgets."""
        from PySide6.QtWidgets import QScrollArea
        
        if not isinstance(scroll_area, QScrollArea):
            return
        
        try:
            # Get and clean the viewport widget
            viewport = scroll_area.viewport()
            if viewport:
                viewport_widget = scroll_area.widget()
                if viewport_widget:
                    self.delete_all_children(viewport_widget)
                    viewport_widget.setParent(None)
                    viewport_widget.deleteLater()
                
                # Clear the scroll area
                scroll_area.setWidget(None)
                
        except (RuntimeError, AttributeError) as e:
            print(f"[ERROR] Error cleaning scroll area: {e}")

    def _close_tab(self, index):
        """Close tab with comprehensive cleanup via CleanupCoordinator.
        
        CRITICAL: Cleanup MUST happen BEFORE any UI operations to prevent
        OCC memory corruption (free(): corrupted unsorted chunks).
        """
        widget = self.tab_widget.widget(index)
        template_instance = self._get_template_instance(index)
        
        # Get module name for debug logging (this may fail for Home tabs)
        module_name = "Unknown"
        try:
            if template_instance and hasattr(template_instance, 'backend') and template_instance.backend:
                module_name = template_instance.backend.module_name()
        except Exception:
            pass
        
        print(f"[TAB CLOSE] Closing tab index {index}: '{module_name}'")
        
        # CRITICAL: Cleanup MUST happen BEFORE UI operations
        # This prevents OCC heap corruption from race conditions
        if template_instance and hasattr(template_instance, 'cad_widget'):
            try:
                # Use AISContextLock if available to prevent race conditions
                try:
                    from osdag_gui.OS_safety_protocols import AISContextLock
                    with AISContextLock():
                        coordinator = get_cleanup_coordinator()
                        coordinator.cleanup_for_tab_close(template_instance)
                except ImportError:
                    # Fallback without lock
                    coordinator = get_cleanup_coordinator()
                    coordinator.cleanup_for_tab_close(template_instance)
            except Exception as e:
                print(f"[TAB CLOSE] Cleanup error: {e}")

        # Remove from UI structures (AFTER cleanup)
        self.tab_widget.removeTab(index)
        self.tab_bar.removeTab(index)
        self.tab_widget_content.pop(index)
        
        # Final widget deletion
        if widget:
            widget.deleteLater()
        
        self._synchronize_tab_widget()
        print(f"[TAB CLOSE] Tab '{module_name}' closed successfully via Coordinator.")
    
    def delete_all_children(self, widget):
            """
            Recursively delete all child widgets of the given widget.
            Traverses depth-first, deleting only QWidget children on the way back up.
            Skips CustomViewer3d to prevent OCC heap corruption.
            """
            from PySide6.QtWidgets import QWidget
            from osdag_gui.ui.components.custom_3dviewer import CustomViewer3d
            
            # Get all immediate children
            children = widget.children()
            
            # Recursively process each child
            for child in children:
                # Only process QWidget instances
                if isinstance(child, QWidget):
                    # Skip CustomViewer3d - deleteLater on it corrupts OCC heap
                    # It will be deleted when parent is deleted
                    if isinstance(child, CustomViewer3d):
                        continue
                    
                    # First, recursively delete this child's children
                    self.delete_all_children(child)
                    
                    # Then delete this child itself
                    child.deleteLater()

    def _synchronize_tab_widget(self):
        """Synchronize tab bar with tab widget content.
        
        CRITICAL: Must handle case when no tabs remain (after last tab closed).
        """
        current_index = self.tab_bar.currentIndex()
        
        # Guard: No tabs left, nothing to synchronize
        if current_index < 0 or len(self.tab_widget_content) == 0:
            return
            
        self.tab_widget.setCurrentIndex(current_index)
        
        # Guard: Index out of bounds
        if current_index >= len(self.tab_widget_content):
            return
            
        # Update global variables and icons
        body_widget = self.tab_widget_content[current_index]
        if hasattr(body_widget, 'layout') and body_widget.layout().count() > 0:
            widget = body_widget.layout().itemAt(0).widget()
            self.main_widget_instance = widget
        # Ensure main_widget_layout points to the currently active tab's layout
        if hasattr(body_widget, 'layout'):
            self.main_widget_layout = body_widget.layout()

    # ============= Resize implementation start ===============
    # MOUSE EVENTS
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            pos = event.pos()
            
            if IS_LINUX:
                # Check if we're in a resize region
                left, right, top, bottom = self.get_resize_region(pos)
                
                # If near any edge, start resize operation
                if left or right or top or bottom:
                    self.resizing = True
                    self.resize_start_pos = event.globalPos()
                    self.resize_start_geometry = self.geometry()
                    
                    # Store which edges we're resizing
                    self.resize_edges = {
                        'left': left,
                        'right': right,
                        'top': top,
                        'bottom': bottom
                    }
                    event.accept()
                    return
                
                # Check if clicking in title bar area
                if pos.y() <= self.title_bar.height():
                    widget = self.childAt(pos)
                    
                    # Don't drag if clicking on buttons
                    if isinstance(widget, QPushButton):
                        super().mousePressEvent(event)
                        return
                    
                    # Check if clicking on an actual tab
                    if widget == self.tab_bar:
                        tab_pos = self.tab_bar.mapFrom(self, pos)
                        tab_index = self.tab_bar.tabAt(tab_pos)
                        if tab_index >= 0:  # Clicking on a tab
                            super().mousePressEvent(event)
                            return
                    
                    # Start dragging
                    self.dragging = True
                    self.drag_position = event.globalPos() - self.frameGeometry().topLeft()
                    event.accept()
                    return
            else:
                # Windows: Only handle dragging for non-maximized state
                if not self.isMaximized():
                    draggable_height = self.tab_bar.height() + (self.layout().contentsMargins().top() * 2)
                    if pos.y() < draggable_height:
                        self.old_pos = event.globalPosition().toPoint()
        
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):
        if IS_LINUX:
            if self.resizing:
                # Handle resizing
                delta = event.globalPos() - self.resize_start_pos
                geo = self.resize_start_geometry
                
                new_x = geo.x()
                new_y = geo.y()
                new_w = geo.width()
                new_h = geo.height()
                
                # Calculate new dimensions based on which edges are being dragged
                if self.resize_edges['left']:
                    new_x = geo.x() + delta.x()
                    new_w = geo.width() - delta.x()
                elif self.resize_edges['right']:
                    new_w = geo.width() + delta.x()
                
                if self.resize_edges['top']:
                    new_y = geo.y() + delta.y()
                    new_h = geo.height() - delta.y()
                elif self.resize_edges['bottom']:
                    new_h = geo.height() + delta.y()
                
                # Apply minimum size constraints
                min_w = self.minimumWidth()
                min_h = self.minimumHeight()
                
                if new_w < min_w:
                    if self.resize_edges['left']:
                        new_x = geo.right() - min_w
                    new_w = min_w
                    
                if new_h < min_h:
                    if self.resize_edges['top']:
                        new_y = geo.bottom() - min_h
                    new_h = min_h
                
                self.setGeometry(new_x, new_y, new_w, new_h)
                event.accept()
                
            elif self.dragging:
                # Handle window dragging
                new_pos = event.globalPos() - self.drag_position
                self.move(new_pos)
                
                # Check for snap preview
                snap_geo = self.get_snap_geometry(event.globalPos())
                self.show_snap_preview(snap_geo)
                
                event.accept()
                
            else:
                # Update cursor when hovering
                self.update_cursor(event.pos())
        else:
            # Windows dragging
            if self.isMaximized():
                return
            if hasattr(self, 'old_pos'):
                delta = event.globalPosition().toPoint() - self.old_pos
                self.move(self.x() + delta.x(), self.y() + delta.y())
                self.old_pos = event.globalPosition().toPoint()
        
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton:
            if IS_LINUX:
                # If we were dragging and there's a snap geometry, apply it
                if self.dragging and self.snap_geometry:
                    # Check if snapping to maximize (full screen)
                    available_screen = QGuiApplication.primaryScreen().availableGeometry()
                    is_maximize_snap = (self.snap_geometry == available_screen)
                    
                    # Save current geometry before snapping (only if not already snapped)
                    if not self.is_snapped and not self.isMaximized():
                        self.pre_snap_geometry = self.geometry()
                    
                    self.setGeometry(self.snap_geometry)
                    
                    # Track snap state
                    self.is_snapped = True
                    self.is_snapped_maximized = is_maximize_snap
                    
                    # Update button icon
                    if is_maximize_snap:
                        self.set_restore_icon()
                    else:
                        self.set_maximize_icon()
                
                self.resizing = False
                self.dragging = False
                self.hide_snap_preview()
            else:
                # Windows
                if hasattr(self, 'old_pos'):
                    del self.old_pos
            
            # restore holding cursor so cursor can update
            self.unsetCursor()
            QApplication.restoreOverrideCursor()
            self.releaseMouse()
            event.accept()
        else:
            super().mouseReleaseEvent(event)

    def mouseDoubleClickEvent(self, event):
        # Toggle maximize/restore when double-clicking in the draggable title area
        if event.button() == Qt.LeftButton:
            pos = event.pos()
            if pos.y() <= TITLEBAR_HEIGHT:
                # Check if we're over an interactive widget
                widget = self.childAt(pos)
                # Only toggle if over empty area, logo, or tab bar empty space
                if (widget is None or 
                    widget == self.icon_label_widget or 
                    widget == self.svg_widget or
                    widget == self.title_bar or
                    widget == self.tab_bar or
                    isinstance(widget, QLabel)):
                    # For tab bar, check we're not on a tab
                    if widget == self.tab_bar:
                        tab_index = self.tab_bar.tabAt(self.tab_bar.mapFromGlobal(event.globalPos()))
                        if tab_index == -1:  # Not over any tab
                            self.toggle_maximize_restore()
                            event.accept()
                            return
                    else:
                        self.toggle_maximize_restore()
                        event.accept()
                        return
        
        super().mouseDoubleClickEvent(event)

    def eventFilter(self, obj, event):
        # Handle double-click on title widgets (e.g., tab bar empty area, logo area)
        if event.type() == QEvent.MouseButtonDblClick:
            if event.button() == Qt.LeftButton:
                # For tab bar, only toggle if not over a tab
                if obj == self.tab_bar:
                    tab_index = self.tab_bar.tabAt(event.pos())
                    if tab_index == -1:  # Not over any tab
                        self.toggle_maximize_restore()
                        return True
                elif obj == self.icon_label_widget or obj == self.svg_widget:
                    self.toggle_maximize_restore()
                    return True
        return super().eventFilter(obj, event)

    def configure_dwm_rendering(self, hwnd, background_color_rgb):
        if IS_WINDOWS:
            margins = ctypes.c_int * 4
            m = margins(-1, -1, -1, -1)
            ctypes.windll.dwmapi.DwmExtendFrameIntoClientArea(hwnd, ctypes.byref(m))
            
            r, g, b = background_color_rgb
            
            # Set dark/light mode
            is_dark = (r + g + b) < 382
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 20, ctypes.byref(ctypes.c_int(1 if is_dark else 0)), 4
            )
            
            # Set frame color (Windows 11)
            colorref = (b << 16) | (g << 8) | r  # BGR format
            try:
                ctypes.windll.dwmapi.DwmSetWindowAttribute(
                    hwnd, 35, ctypes.byref(ctypes.c_int(colorref)), 4
                )
            except: pass

    # Show window cleanly with proper style and shadow
    def show(self):
        if IS_WINDOWS:
            hwnd = int(self.winId())
            apply_window_style(hwnd)
            bg_rgb = (255, 255, 255)  # BG Color for Windows Wrapper
            self.configure_dwm_rendering(hwnd, bg_rgb)
            self.setAttribute(Qt.WA_DontShowOnScreen, False)
        
        super().show()

    # ============= Resize implementation ends ===============

    def handle_card_open_clicked(self, card_title):
        # print(f"[INFO] Card opened: {card_title}")

        #----------Shear-Connections--------------
        if card_title == KEY_DISP_FINPLATE:
            self.open_fin_plate_shear_conn()
        elif card_title == KEY_DISP_CLEATANGLE:
            self.open_cleat_angle_shear_conn()
        elif card_title == KEY_DISP_ENDPLATE:
            self.open_header_plate_shear_conn()
        elif card_title == KEY_DISP_SEATED_ANGLE:
            self.open_seated_angle_shear_conn()

        #----------Beam-to-Column-Connections--------------
        elif card_title == KEY_DISP_BCENDPLATE:
            self.open_btc_end_plate_moment_conn() 

        #----------Beam-to-Beam-Connections--------------
        elif card_title == KEY_DISP_BEAMCOVERPLATEWELD:
            self.open_btb_cover_plate_weld_moment_conn()
        elif card_title == KEY_DISP_BEAMCOVERPLATE:
            self.open_btb_cover_plate_bolt_moment_conn()
        elif card_title == KEY_DISP_BB_EP_SPLICE:
            self.open_btb_end_plate_moment_conn()

        #----------Column-to-Column-Connections--------------
        elif card_title == KEY_DISP_COLUMNCOVERPLATE:
            self.open_ctc_cover_plate_bolt_moment_conn()
        elif card_title == KEY_DISP_COLUMNCOVERPLATEWELD:
            self.open_ctc_cover_plate_weld_moment_conn()
        elif card_title == KEY_DISP_COLUMNENDPLATE:
            self.open_ctc_end_plate_moment_connection()

        #----------Simple-Connections--------------
        elif card_title == KEY_DISP_LAPJOINTWELDED:
            self.open_lap_joint_welded_simple_conn()
        elif card_title == KEY_DISP_LAPJOINTBOLTED:
            self.open_lap_joint_bolted_simple_conn()
        elif card_title == KEY_DISP_BUTTJOINTBOLTED:
            self.open_butt_joint_bolted_simple_conn()
        elif card_title == KEY_DISP_BUTTJOINTWELDED:
            self.open_butt_joint_welded_simple_conn()

        #----------Tension-Member--------------
        elif card_title == KEY_DISP_TENSION_BOLTED:
            self.open_tension_bolted()
        elif card_title == KEY_DISP_TENSION_WELDED:
            self.open_tension_welded()

        #----------Compression-Member--------------
        elif card_title == KEY_DISP_COMPRESSION_COLUMN:
            self.open_column_design_compress_member()
        elif card_title == KEY_DISP_STRUT_WELDED_END_GUSSET:
            self.open_struts_weld_end_gusset_compress_member()
        elif card_title == KEY_DISP_STRUT_BOLTED_END_GUSSET:
            self.open_struts_bolted_end_gusset_compress_member()

        #----------Flexure-Member--------------
        elif card_title == KEY_DISP_FLEXURE:
            self.open_simply_supported_beam_flexure()
        elif card_title == KEY_DISP_FLEXURE2:
            self.open_cantilever_beam_flexure()
        elif card_title == KEY_DISP_PLATE_GIRDER_WELDED:
            self.open_plate_girder_flexure()
        elif card_title == KEY_DISP_FLEXURE4:
            self.open_purlin_flexure()

        #---------Base Plate Connection------------------
        elif card_title == KEY_DISP_BASE_PLATE:
            self.open_base_plate_conn()

    # TO update count of opened module in current session
    def update_module_count(self, backend:object) -> int:
        key = backend.module_name()
        # Increment module count
        self.module_count[key] += 1
        return self.module_count[key]

    #-------------Functions-to-load-modules-in-Tabwidget-START---------------------------
    def common_open_module(self, backend_class, title, id):
        self.clear_layout(self.main_widget_layout)
        template_page = CustomWindow(title, backend_class, id, parent=self)

        template_page.setWindowFlags(Qt.Widget)
        template_page.setAttribute(Qt.WA_DontCreateNativeAncestors, True)
        template_page.setAttribute(Qt.WA_NativeWindow, False)
        
        # Prevent all children from creating native windows
        # IMPORTANT: This enables event detection after opening template_page
        for child in template_page.findChildren(QWidget):
            child.setAttribute(Qt.WA_DontCreateNativeAncestors, True)

        # Load the last Design Inputs-start------------------------------------
        last_design_folder = os.path.join('ResourceFiles', 'last_designs')
        last_design_file = str(template_page.backend.module_name()).replace(' ', '') + ".osi"
        last_design_file = os.path.join(last_design_folder, last_design_file)
        last_design_dictionary = {}

        # Create folder if it doesn't exist
        if not os.path.isdir(last_design_folder):
            os.makedirs(last_design_folder)

        # Load previous design if file exists
        if os.path.isfile(last_design_file):
            with open(str(last_design_file), 'r') as last_design:
                last_design_dictionary = yaml.safe_load(last_design)
                template_page.setDictToUserInputs(last_design_dictionary)
        # Load the last Design Inputs-end------------------------------------

        self.main_widget_instance = template_page
        template_page.openNewTab.connect(self.handle_add_tab)
        template_page.downloadDatabase.connect(self.download_Database)
        self.main_widget_layout.addWidget(template_page)
        
        index = self.tab_bar.currentIndex()
        self.tab_bar.setTabText(index, title)
    
    # 1-Fin-plate-shear-connection
    def open_fin_plate_shear_conn(self):
        # Local import to avoid empty material list in UI
        from osdag_core.design_type.connection.fin_plate_connection import FinPlateConnection
        backend = FinPlateConnection
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Fin Plate Connection", id)

    # 2-Cleat-angle-shear-connection
    def open_cleat_angle_shear_conn(self):
        from osdag_core.design_type.connection.cleat_angle_connection import CleatAngleConnection
        backend = CleatAngleConnection
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Cleat Angle Connection", id)

    # 3-Header-plate-shear-connection
    def open_header_plate_shear_conn(self):
        from osdag_core.design_type.connection.end_plate_connection import EndPlateConnection
        backend = EndPlateConnection
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Header Plate Connection", id)  
    
    # 4-Seated-angle-shear-connection
    def open_seated_angle_shear_conn(self):
        from osdag_core.design_type.connection.seated_angle_connection import SeatedAngleConnection
        backend = SeatedAngleConnection
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Seated Angle Connection", id)
    
    # 5-Beam-to-Column-end-plate-moment-connection
    def open_btc_end_plate_moment_conn(self):
        from osdag_core.design_type.connection.beam_column_end_plate import BeamColumnEndPlate
        backend = BeamColumnEndPlate
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Beam Column End Plate Connection", id)

    # 6-Beam-to-Beam-cover-plate-welded-moment-connection
    def open_btb_cover_plate_weld_moment_conn(self):
        from osdag_core.design_type.connection.beam_cover_plate_weld import BeamCoverPlateWeld
        backend = BeamCoverPlateWeld
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Beam Beam Cover Plate Welded", id)

    # 7-Beam-to-Beam-cover-plate-bolted-moment-connection
    def open_btb_cover_plate_bolt_moment_conn(self):
        from osdag_core.design_type.connection.beam_cover_plate import BeamCoverPlate
        backend = BeamCoverPlate
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Beam Beam Cover Plate Bolted", id)
        
    # 8-Beam-to-Beam-end-plate-splice-moment-connection
    def open_btb_end_plate_moment_conn(self):
        from osdag_core.design_type.connection.beam_beam_end_plate_splice import BeamBeamEndPlateSplice
        backend = BeamBeamEndPlateSplice
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Beam Beam End Plate", id)

    # 9-Column-to-Column-end-plate-moment-connection
    def open_ctc_end_plate_moment_connection(self):
        from osdag_core.design_type.connection.column_end_plate import ColumnEndPlate
        backend = ColumnEndPlate
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Column End plate", id)    

    # 10-Column-to-Column-cover-plate-bolted-moment-connection
    def open_ctc_cover_plate_bolt_moment_conn(self):
        from osdag_core.design_type.connection.column_cover_plate import ColumnCoverPlate
        backend = ColumnCoverPlate
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Column Cover Plate Bolted", id)
        
    # 11-Column-to-Column-cover-plate-welded-moment-connection
    def open_ctc_cover_plate_weld_moment_conn(self):
        from osdag_core.design_type.connection.column_cover_plate_weld import ColumnCoverPlateWeld
        backend = ColumnCoverPlateWeld
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Column Cover Plate Welded", id)

    # 12-Lap-joint-welded-simple-Connection
    def open_lap_joint_welded_simple_conn(self):
        from osdag_core.design_type.connection.lap_joint_welded import LapJointWelded
        backend = LapJointWelded
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Lap Joint Welded Connection", id)

    # 13-Lap-joint-bolted-simple-connection
    def open_lap_joint_bolted_simple_conn(self):
        from osdag_core.design_type.connection.lap_joint_bolted import LapJointBolted
        backend = LapJointBolted
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Lap Joint Bolted Connection", id)
        
    # 14-Butt-joint-bolted-simple-connection
    def open_butt_joint_bolted_simple_conn(self):
        from osdag_core.design_type.connection.butt_joint_bolted import ButtJointBolted
        backend = ButtJointBolted
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Butt Joint Bolted Connection", id)

    # 15-Butt-joint-welded-simple-connection
    def open_butt_joint_welded_simple_conn(self):
        from osdag_core.design_type.connection.butt_joint_welded import ButtJointWelded
        backend = ButtJointWelded
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Butt Joint Welded Connection", id) 

    # 16-Bolted-to-End-Gusset-Tension-Member
    def open_tension_bolted(self):
        from osdag_core.design_type.tension_member.tension_bolted import Tension_bolted
        backend = Tension_bolted
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Tension Member: Bolted to End Gusset", id)
 
    # 17-Welded-to-End-Gusset-Tension-Member
    def open_tension_welded(self):
        from osdag_core.design_type.tension_member.tension_welded import Tension_welded
        backend = Tension_welded
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Tension Member: Welded to End Gusset", id)     
 
    # 18-Column-design-Compression-Member
    def open_column_design_compress_member(self):
        from osdag_core.design_type.compression_member.compression_column import ColumnDesign
        backend = ColumnDesign
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Column Design", id)

    # 19-Struts-welded-to-end-gusset-compression-member
    def open_struts_weld_end_gusset_compress_member(self):
        from osdag_core.design_type.compression_member.compression_welded import Compression_welded
        backend = Compression_welded
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Struts: Welded to End Gusset", id)

    def open_struts_bolted_end_gusset_compress_member(self):
        from osdag_core.design_type.compression_member.compression_bolted import Compression_bolted
        backend = Compression_bolted
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Struts: Bolted to End Gusset", id)

    # 20-Simply-Supported-Beam-Flexure-member
    def open_simply_supported_beam_flexure(self):
        from osdag_core.design_type.flexural_member.flexure import Flexure
        backend = Flexure
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Simply Supported Beam", id)
        
    # 21-Cantilever-Beam-Flexure-member
    def open_cantilever_beam_flexure(self):
        from osdag_core.design_type.flexural_member.flexure_cantilever import Flexure_Cantilever
        backend = Flexure_Cantilever
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Cantilever Beam", id)

    # 22-Plate-girder
    def open_plate_girder_flexure(self):
        from osdag_core.design_type.plate_girder.weldedPlateGirder import PlateGirderWelded
        backend = PlateGirderWelded
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Plate Girder", id)  

    # 23-Flexure-purlin
    def open_purlin_flexure(self):
        from osdag_core.design_type.flexural_member.flexure_purlin import Flexure_Purlin
        backend = Flexure_Purlin
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Purlin", id)

    # 24-Base-Plate-connection
    def open_base_plate_conn(self):
        from osdag_core.design_type.connection.base_plate_connection import BasePlateConnection
        backend = BasePlateConnection
        id = self.update_module_count(backend)
        self.common_open_module(backend, "Base Plate Connection", id)

    def open_home_page(self, module):
        self.clear_layout(self.main_widget_layout)
        home_window = HomeWindow()
        home_window.triggerLoadOsi.connect(self.common_osi_load)
        home_window.openProject.connect(self.handle_open_project)
        home_window.openModule.connect(self.handle_open_module)
        home_window.downloadDatabase.connect(self.download_Database)
        self.main_widget_instance = home_window
        home_window.set_active_button(module)
        home_window.cardOpenClicked.connect(self.handle_card_open_clicked)
        self.main_widget_layout.addWidget(home_window)

    # To open the recent module
    def handle_open_module(self, key:str):
        func = get_module_function(key)
        if func != 'None':
            func = getattr(self, func)
            func() # Open the Releated Module

    # To handle the click on open project of any recent project
    def handle_open_project(self, record: dict):
        self.common_osi_load(osi_path=record.get(PROJECT_PATH), id=record.get(ID))

    # Common function to load osi file and also to open recent project
    # If osi_path=None -> it triggers Load Osi else trigger open recent project
    def common_osi_load(self, osi_path=None, id=None):
        if osi_path is None:
            osi_path, _ = QFileDialog.getOpenFileName(self, "Open Design", os.path.join(str(' ')),
                                                  "InputFiles(*.osi)")
            
        else:
            if not Path(osi_path).exists():
                result = CustomMessageBox(
                    title="Warning",
                    text="Osi File has been moved, File does not exist!",
                    dialogType=MessageBoxType.Warning,
                    buttons=["Locate Osi", "Remove Record"]
                ).exec()
                if result == "Locate Osi":
                    file_dialog_path, _ = QFileDialog.getOpenFileName(self, "Locate Osi File", os.path.expanduser("~"), "InputFiles(*.osi)")
                    if file_dialog_path and id is not None:
                        osi_path = file_dialog_path
                        new_name = Path(osi_path).stem
                        try:
                            update_project_path(id, osi_path, new_name)
                        except Exception as e:
                            print(f"[ERROR] Failed to update project path: {e}")
                    else:
                        print("[INFO] No file selected for relocation.")
                        return
                elif result == "Remove Record":
                    print("[INFO] Remove Record")
                    if id is not None:
                        try:
                            delete_project_record(id)
                            # Also delete the latex files for this project
                            import shutil
                            report_folder = f"osdag_gui/data/reports/file_{id}"
                            try:
                                shutil.rmtree(report_folder)
                            except FileNotFoundError:
                                pass
                            except Exception as e:
                                print(f"[ERROR] Failed to delete report folder: {e}")
                            CustomMessageBox(
                                title="Record Removed",
                                text="The record has been removed from recent projects.",
                                dialogType=MessageBoxType.Information
                            ).exec()
                            # Update Home page with deleted project
                            self.main_widget_instance.show_home()

                        except Exception as e:
                            CustomMessageBox(
                                title="Error",
                                text=f"Failed to remove record: {e}",
                                dialogType=MessageBoxType.Critical
                            ).exec()
                    else:
                        print("[INFO] No ID provided for record removal.")
                    return

        if not osi_path:
            print("[INFO] No Path selected!")
            return
        try:
            in_file = str(osi_path)
            with open(in_file, 'r') as fileObject:
                uiObj = yaml.safe_load(fileObject)
            module = uiObj[KEY_MODULE]

            print(f"[INFO] Osi File Belongs to: {module}")

            func = get_module_function(module)
            if func == 'None':
                CustomMessageBox(
                    title="Information",
                    text="Please load the appropriate Input",
                    dialogType=MessageBoxType.Information
                ).exec()
                print("[INFO] Module Under Development.")
                return
            func = getattr(self, func)
            func()
            # Set variables in template page because it is opened project
            self.main_widget_instance.setDictToUserInputs(uiObj)
            self.main_widget_instance.project_id = id
            self.main_widget_instance.save_state = True

        except IOError:
            CustomMessageBox(
                title="Unable to open file",
                text="There was an error opening \"%s\"" % osi_path,
                dialogType=MessageBoxType.Critical
            ).exec()
            return    

    #-------------Functions-to-load-modules-in-Tabwidget-END------------------------------------

    #----------------------------Download-Database/Excel-END-----------------------------------------
    def download_Database(self, table, call_type="database"):

        default_dir = os.path.join(get_documents_folder(), str(table+"_Details.xlsx"))
        fileName, _ = QFileDialog.getSaveFileName(  QFileDialog(), 
                                                    "Download File",
                                                    default_dir,
                                                    "SectionDetails(*.xlsx)"
                                                )
        if not fileName:
            return
        try:
            import sqlite3
            conn = sqlite3.connect(PATH_TO_DATABASE)
            c = conn.cursor()
            header = get_db_header(table)
            wb = openpyxl.Workbook()
            sheet = wb.create_sheet(table, 0)

            col = 1
            for head in header:
                sheet.cell(row=1, column=col).value = head
                col += 1
            if call_type != "header":
                if table == 'Columns':
                    c.execute("SELECT * FROM Columns")
                elif table == 'Beams':
                    c.execute("SELECT * FROM Beams")
                elif table == 'Angles':
                    c.execute("SELECT * FROM Angles")
                elif table == 'Channels':
                    c.execute("SELECT * FROM Channels")
                data = c.fetchall()
                conn.commit()
                c.close()
                row = 2
                for rows in data:
                    col = 1
                    for cols in range(len(header)):
                        sheet.cell(row=row, column=col).value = rows[col - 1]
                        col += 1
                    row += 1
            wb.save(fileName)
            CustomMessageBox(
                title='Information',
                text='Your File is Downloaded.',
                dialogType=MessageBoxType.Information
            ).exec()

        except IOError:
            CustomMessageBox(
                title='Information',
                text='Unable to save file',
                informativeText="There was an error saving \"%s\"" % fileName,
                dialogType=MessageBoxType.Information
            ).exec()
            return
        
    def closeEvent(self, event):
        """Clean up any remaining tabs and exit.
        
        Since tabs are closed individually by user in a user-led flow,
        this is just a fallback for alt-F4 or window manager close.
        """
        # If tabs remain (user closed via alt-F4 or window manager)
        if hasattr(self, 'tab_widget_content') and len(self.tab_widget_content) > 0:
            print(f"[APP EXIT] Closing {len(self.tab_widget_content)} remaining tabs")
            self._closing_tabs = True
            # Close tabs one by one - this uses existing _close_tab cleanup
            while self.tab_bar.count() > 0:
                self._close_tab(0)
        
        event.accept()
        self.deleteLater()

    #----------------------------Download-Database/Excel--END----------------------------------------

# if __name__ == "__main__":
#     import sys, os
#     from osdag_gui.ui.utils.theme_manager import ThemeManager
#     sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
#     from PySide6.QtWidgets import QApplication
#     app = QApplication(sys.argv)
#     app.theme_manager = ThemeManager(app)
#     main_window = MainWindow()
#     main_window.show()
#     sys.exit(app.exec())
    
    
    
